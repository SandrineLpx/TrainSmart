"""
Parse a training program Excel file into data/program.json.

Usage:
  python scripts/parse_excel.py                          # uses default "Leg Drive.xlsx"
  python scripts/parse_excel.py "My Program.xlsx"        # use a different file
  python scripts/parse_excel.py program.xlsx 2026-03-01  # custom file + start date
  python scripts/parse_excel.py --name "Squat Cycle"     # custom program name

Expected Excel format:
  - One sheet per week (named anything: "Week 1", "Phase 1", etc.)
  - Columns: A=Day, B=Exercise, C=Sets, D=Reps, E=Notes
  - Day names in column A (Monday, Tuesday, etc.) as section headers
  - Exercises listed under each day header
"""

import argparse
import json
import os
import sys

import openpyxl

PROJECT_ROOT = os.path.join(os.path.dirname(__file__), "..")
OUTPUT_PATH = os.path.join(PROJECT_ROOT, "data", "program.json")

VALID_DAYS = {"Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"}


def parse_sheet(ws):
    """Parse a single week sheet into a dict of {day: [exercises]}."""
    days = {}
    current_day = None

    for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
        col_a = row[0]  # DAY
        col_b = row[1]  # EXERCISE
        col_c = row[2]  # SETS
        col_d = row[3]  # REPS
        col_e = row[4]  # NOTES

        # Strip strings
        if isinstance(col_a, str):
            col_a = col_a.strip()
        if isinstance(col_b, str):
            col_b = col_b.strip()
        if isinstance(col_e, str):
            col_e = col_e.strip()

        # New day header
        if col_a and col_a in VALID_DAYS:
            current_day = col_a
            if current_day not in days:
                days[current_day] = []
            if col_b:
                days[current_day].append({
                    "exercise": col_b,
                    "sets": _to_val(col_c),
                    "reps": _to_val(col_d),
                    "notes": col_e if col_e else None,
                })
            continue

        # Exercise row
        if current_day and col_b:
            days[current_day].append({
                "exercise": col_b,
                "sets": _to_val(col_c),
                "reps": _to_val(col_d),
                "notes": col_e if col_e else None,
            })

    return days


def _to_val(v):
    """Convert cell value to a clean string or int."""
    if v is None:
        return None
    if isinstance(v, float) and v == int(v):
        return int(v)
    if isinstance(v, str):
        return v.strip()
    return v


def main():
    parser = argparse.ArgumentParser(
        description="Parse a training program Excel file into data/program.json"
    )
    parser.add_argument(
        "excel_file", nargs="?", default=os.path.join(PROJECT_ROOT, "references", "Leg Drive.xlsx"),
        help="Path to the Excel file (default: Leg Drive.xlsx)"
    )
    parser.add_argument(
        "start_date", nargs="?", default=None,
        help="Program start date YYYY-MM-DD (default: asks interactively)"
    )
    parser.add_argument(
        "--name", default=None,
        help="Program name (default: filename without extension)"
    )
    args = parser.parse_args()

    # Validate file exists
    if not os.path.exists(args.excel_file):
        print(f"ERROR: File not found: {args.excel_file}")
        sys.exit(1)

    # Program name: flag > filename
    if args.name:
        program_name = args.name
    else:
        program_name = os.path.splitext(os.path.basename(args.excel_file))[0]

    # Start date: argument > interactive
    if args.start_date:
        start_date = args.start_date
    else:
        start_date = input(f"Program start date (YYYY-MM-DD): ").strip()
        if not start_date:
            print("ERROR: Start date is required.")
            sys.exit(1)

    # Parse
    wb = openpyxl.load_workbook(args.excel_file, read_only=True, data_only=True)

    program = {
        "program_name": program_name,
        "program_start_date": start_date,
        "total_weeks": len(wb.sheetnames),
        "weeks": [],
    }

    for i, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        week_num = i + 1
        days = parse_sheet(ws)

        program["weeks"].append({
            "week_number": week_num,
            "pattern": "odd" if week_num % 2 == 1 else "even",
            "days": days,
        })

    wb.close()

    # Also update the start date in CLAUDE.md and preferences.json
    _update_start_date(start_date)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(program, f, indent=2, ensure_ascii=False)

    print(f"\nProgram loaded: {program_name}")
    print(f"Start date: {start_date}")
    print(f"Saved to: {OUTPUT_PATH}")
    print(f"Weeks: {program['total_weeks']}")
    for w in program["weeks"]:
        day_summary = ", ".join(
            f"{d}: {len(exs)} ex" for d, exs in w["days"].items()
        )
        print(f"  Week {w['week_number']} ({w['pattern']}): {day_summary}")

    print("\nYou can also edit data/program.json directly at any time.")

    # Generate readable summary
    summary_path = _write_summary(program)
    print(f"Summary saved to: {summary_path}")


def _write_summary(program):
    """Generate a human-readable markdown summary of the program."""
    name = program["program_name"]
    start = program["program_start_date"]
    total = program["total_weeks"]
    summary_path = os.path.join(PROJECT_ROOT, "data", "program_summary.md")

    lines = [
        f"# {name} — Program Summary",
        "",
        f"**Start date:** {start}  ",
        f"**Duration:** {total} weeks  ",
        f"**Days per week:** {len(program['weeks'][0]['days'])} prescribed  ",
        "",
        "---",
        "",
    ]

    for week in program["weeks"]:
        wn = week["week_number"]
        pattern = week["pattern"]
        lines.append(f"## Week {wn} ({pattern})")
        lines.append("")

        day_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        for day in day_order:
            if day not in week["days"]:
                continue
            exercises = week["days"][day]
            lines.append(f"### {day}")
            for ex in exercises:
                sets = ex["sets"] if ex["sets"] is not None else "—"
                reps = ex["reps"] if ex["reps"] is not None else "—"
                entry = f"- **{ex['exercise']}** — {sets} x {reps}"
                if ex.get("notes"):
                    entry += f"  *({ex['notes']})*"
                lines.append(entry)
            lines.append("")

    # Volume tracking across weeks
    lines.append("---")
    lines.append("")
    lines.append("## Volume Progression")
    lines.append("")
    lines.append("| Week | Pattern | Main lift sets | Days |")
    lines.append("|------|---------|---------------|------|")
    for week in program["weeks"]:
        wn = week["week_number"]
        pattern = week["pattern"]
        days = list(week["days"].keys())
        # Find the first exercise with sets defined as a proxy for main lift volume
        main_sets = "—"
        for day_exs in week["days"].values():
            for ex in day_exs:
                if ex["sets"] is not None and isinstance(ex["sets"], int) and ex["sets"] >= 3:
                    main_sets = str(ex["sets"])
                    break
            if main_sets != "—":
                break
        lines.append(f"| {wn} | {pattern} | {main_sets} | {', '.join(d[:3] for d in days)} |")

    lines.append("")

    with open(summary_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    return summary_path


def _update_start_date(start_date):
    """Update start date in preferences.json."""
    prefs_path = os.path.join(PROJECT_ROOT, "data", "preferences.json")
    try:
        with open(prefs_path, encoding="utf-8") as f:
            prefs = json.load(f)
        prefs["program_start_date"] = start_date
        with open(prefs_path, "w", encoding="utf-8") as f:
            json.dump(prefs, f, indent=2)
    except FileNotFoundError:
        pass


if __name__ == "__main__":
    main()
